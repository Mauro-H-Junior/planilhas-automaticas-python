import openpyxl
import time

# -----------------------------------------------------------

colunas = []

planilha_python = openpyxl.Workbook()
planilha_python.remove(planilha_python['Sheet'])

# -----------------------------------------------------------

def imprimir_titulos(mensagem):
    print(87*'-')
    print(mensagem)
    print(87*'-')
    time.sleep(2)

def verificar_escolha_do_usuario_sim_nao(mensagem):
    sim_ou_nao = input(f'\n{mensagem}\nDigite "sim" ou "não":')
    return sim_ou_nao.lower() == 'sim'

def criar_planilhas():
    while True:
        nome_da_planilha = input ('\n***** Qual o nome da planilha/aba que deseja criar? ***** \nNome da planilha/aba:')
        print(87*'-')
        planilha_python.create_sheet(nome_da_planilha)

        continuar = verificar_escolha_do_usuario_sim_nao('Deseja criar uma nova aba/planilha?')
        if not continuar:
            break

def exibir_planilhas():

    print('\n***** Abaixo irei listar as abas/planilhas que você criou: *****')
    time.sleep(2)
    print('')
    print(87*'-')
    print('Listando....')
    time.sleep(2)

    for i in planilha_python.sheetnames:
        print(f'Você tem a planilha: {i}')
        print(87*'-')

def sair_salvar_trabalho():    
 
    continuar = verificar_escolha_do_usuario_sim_nao('***** Deseja CONTINUAR NO SISTEMA? *****')
    if not continuar:
        
        print(87*'-')
        nome_do_projeto = input('***** Tudo bem vou salvar e encerrar o APP. Qual será o nome do seu projeto em Excel? *****\nNome do projeto: ')
        planilha_python.save(f'{nome_do_projeto}.xlsx')
        print('Finalizando APP..')
        time.sleep(2)
        return True
    return False

def itera_existencia_planilha_na_lista(planilha):
    for i in planilha_python.sheetnames:
        if i == planilha:
            return True
    return False

def verifica_existencia_e_seleciona_planilhas():

    while True:
        aba_para_modificacao = input('\nCerto, qual das abas/planilhas acima você deseja modificar?')
        
        if itera_existencia_planilha_na_lista(aba_para_modificacao):
            return planilha_python[aba_para_modificacao] 

        print('\n***** Favor, digite exatamente o nome da planilha *****')
        time.sleep(2)
        print(87*'-')
        exibir_planilhas()    

def criar_colunas(pagina_para_alteracao):
    while True:

        print('Preparando aba/planilha...')
        time.sleep(2)
        print(87*'-')
        nome_do_cabecalho = input('\n***** OK! Qual o NOME DA COLUNA que deseja inserir? *****\nNome da coluna:')
        print(87*'-')
        colunas.append(nome_do_cabecalho)
    
        continuar = verificar_escolha_do_usuario_sim_nao('\n***** Deseja criar MAIS UMA COLUNA no cabeçalho? *****')
        if not continuar:
            pagina_para_alteracao.append(colunas)
            colunas.clear()
            break
        
def criar_linhas(pagina_para_alteracao):
    while True:

        nome_da_nova_linha = input('\n***** Para incluir uma linha digite os valores separados por vírgulas.*****\nExemplo: 12, RAFAEL, TECNICO:')
        pagina_para_alteracao.append(nome_da_nova_linha.split(','))
       
    
        continuar = verificar_escolha_do_usuario_sim_nao('\n***** Deseja criar MAIS UMA LINHA na tabela? *****')
        if not continuar:
            # pagina_para_alteracao.append(linhas)
            # linhas.clear()
            break

def modificacao_de_outra_planilha():
    while True:
        continuar = verificar_escolha_do_usuario_sim_nao('***** Certo! Deseja MODIFICAR OUTRA aba/planilha? *****')
        time.sleep(2)
        if continuar:
            exibir_planilhas()
            seleciona_planilha_para_alteracao = verifica_existencia_e_seleciona_planilhas()
            criar_colunas(seleciona_planilha_para_alteracao)
            return False
        return True

def iterar_linhas_da_planilha():
    for coluna in seleciona_planilha_para_alteracao.iter_rows(min_row=1,max_row=1):
        for row in coluna: #Aqui eu itero para cada célula nessa linha
            print(87*'-') 
            if row.value == None:
                print('Não há colunas nessa planilha')
            print(f'Coluna dessa aba/planilha:{row.value}')
            print(87*'-') 

def incluir_nova_linha_planilha():

    continuar = verificar_escolha_do_usuario_sim_nao('***** Certo, agora deseja incluir uma nova linha em alguma das planilhas criadas? *****')
    if continuar:
        exibir_planilhas()
        seleciona_planilha_para_alteracao = verifica_existencia_e_seleciona_planilhas()
        

        print(f'\n***** Ok! Você SELECIONOU a planilha/aba: {seleciona_planilha_para_alteracao}. Abaixo irei listar quais as colunas dessa planilha *****')
        print('Listando...')
        time.sleep(2)

        iterar_linhas_da_planilha()
    
        criar_linhas(seleciona_planilha_para_alteracao)

        return False
    return True

#------------------------------------------------

imprimir_titulos('***** BEM VINDO(A) AO APP DE CRIAÇÃO DE ARQUIVOS EXCEL COM PYTHON *****')
criar_planilhas()

while True:
    exibir_planilhas()
    if sair_salvar_trabalho():
        break
    seleciona_planilha_para_alteracao = verifica_existencia_e_seleciona_planilhas()
    criar_colunas(seleciona_planilha_para_alteracao)
    incluir_nova_linha_planilha()
    if sair_salvar_trabalho():
        break



 



            